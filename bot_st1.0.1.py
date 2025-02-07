# 导入必要的库
import streamlit as st  # Web应用框架
import traceback  # 用于详细的错误追踪
from docx import Document  # 处理Word文档
from openpyxl import Workbook  # 处理Excel文件
import re  # 正则表达式处理
import io  # 处理内存中的字节流
import base64  # 用于数据编码

# Streamlit页面配置（必须在其他Streamlit命令之前）
st.set_page_config(
    page_title="Word试题转Excel工具",  # 设置页面标题
    page_icon="📚",  # 设置页面图标
    layout="wide",  # 使用宽屏布局
    initial_sidebar_state="expanded"  # 侧边栏默认展开
)

def extract_questions_from_docx(doc):
    """
    从Word文档中提取题目信息
    参数:
        doc: Document对象，包含题目内容
    返回:
        questions: 列表，包含所有题目信息的字典
    """
    questions = []  # 存储所有题目
    current_question = {}  # 当前正在处理的题目
    question_type = ''  # 当前题目类型
    
    for para in doc.paragraphs:
        text = para.text.strip()  # 去除首尾空白
        if not text:
            continue
            
        # 处理题型标记（如 ##单选题##）
        if text.startswith('##') and text.endswith('##'):
            question_type = text.strip('#').strip()
            continue
            
        # 处理题目（以数字开头的段落）- 修改正则表达式以适应更多格式
        if re.match(r'^[#]?\d+\.?\s', text):  # 添加可选的点号
            if current_question:  # 如果存在上一题，保存它
                questions.append(current_question)
            # 初始化新题目
            current_question = {
                'type': question_type,
                'question': '',
                'options': [],
                'answer': ''
            }
            
            # 提取题目内容和答案
            question_text = re.sub(r'^[#]?\d+\.?\s', '', text).strip()  # 移除题号和可能的点号
            # 提取答案（在{}中的内容）
            answer_match = re.search(r'\{(.+?)\}', question_text)
            if answer_match:
                current_question['answer'] = answer_match.group(1)
                question_text = re.sub(r'\{.+?\}', '', question_text)  # 移除答案标记
            current_question['question'] = question_text.strip()
            
        # 处理选项（A-D开头的行）- 修改以适应带点号的格式
        elif re.match(r'^[A-D]\.?\s', text):  # 添加可选的点号
            option_text = re.sub(r'^[A-D]\.?\s', '', text).strip()  # 移除选项标记和可能的点号
            current_question['options'].append(option_text)
    
    # 保存最后一题
    if current_question:
        questions.append(current_question)
    
    return questions

def create_excel(questions):
    """
    将题目信息转换为Excel文件
    参数:
        questions: 包含题目信息的列表
    返回:
        excel_buffer: 包含Excel文件内容的字节流
    """
    wb = Workbook()
    ws = wb.active
    
    # 设置表头
    headers = ['题目类型', '题目', 'A', 'B', 'C', 'D', '正确答案']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 填充题目数据
    current_row = 2
    for q in questions:
        ws.cell(row=current_row, column=1, value=q['type'])
        ws.cell(row=current_row, column=2, value=q['question'])
        
        # 填充选项（仅针对单选题和多选题）
        if q['type'] in ['单选题', '多选题']:
            for i, option in enumerate(q['options']):
                ws.cell(row=current_row, column=i+3, value=option)
        
        ws.cell(row=current_row, column=7, value=q['answer'])
        current_row += 1
    
    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # 将Excel文件保存到内存中
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

def main():
    """主函数：处理用户界面和程序流程"""
    # 设置页面标题
    st.title("Word试题转Excel工具")
    st.write("支持的题型：单选题、多选题、判断题")
    
    # 显示使用说明（可折叠）
    with st.expander("查看使用说明"):
        st.markdown("""
        ### 使用说明：
        1. 上传Word文档（.docx格式）
        2. 文档格式要求：
           - 题型标记使用 ##单选题##、##多选题##、##判断题## 格式
           - 每道题以数字开头
           - 答案放在题目末尾的大括号中，如：{A}
           - 选项以A、B、C、D开头
        """)
    
    # 文件上传组件
    uploaded_file = st.file_uploader("选择Word文档", type=['docx'])
    
    # 处理上传的文件
    if uploaded_file:
        try:
            # 读取和处理Word文档
            doc = Document(uploaded_file)
            questions = extract_questions_from_docx(doc)
            
            if questions:
                # 生成Excel文件
                excel_buffer = create_excel(questions)
                
                # 创建下载按钮
                st.download_button(
                    label="下载转换后的Excel文件",
                    data=excel_buffer.getvalue(),
                    file_name="转换结果.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # 显示转换结果预览
                st.subheader("转换结果预览")
                for i, q in enumerate(questions, 1):
                    st.write(f"**{i}. {q['type']}**")
                    st.write(f"题目：{q['question']}")
                    if q['options']:
                        st.write("选项：")
                        for j, opt in enumerate(q['options']):
                            st.write(f"{chr(65+j)}. {opt}")
                    st.write(f"答案：{q['answer']}")
                    st.write("---")
                
        except Exception as e:
            st.error(f"处理文件时出错：{str(e)}")
                
# 程序入口点
if __name__ == '__main__':
    main()