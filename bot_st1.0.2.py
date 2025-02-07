# 导入必要的库
import streamlit as st  # Web应用框架
import traceback  # 用于详细的错误追踪
from docx import Document  # 处理Word文档
from openpyxl import Workbook  # 处理Excel文件
import re  # 正则表达式处理
import io  # 处理内存中的字节流
import base64  # 用于数据编码
import os  # 用于文件操作

# Streamlit页面配置（必须在其他Streamlit命令之前）
st.set_page_config(
    page_title="Word试题转Excel工具",  # 设置页面标题
    page_icon="📚",  # 设置页面图标
    layout="wide",  # 使用宽屏布局
    initial_sidebar_state="expanded"  # 侧边栏默认展开
)

def extract_questions_from_docx(doc):
    """
    从Word文档中提取题目信息
    参数:
        doc: Document对象，包含题目内容
    返回:
        questions: 列表，包含所有题目信息的字典
    """
    questions = []  # 存储所有题目
    current_question = {}  # 当前正在处理的题目
    question_type = ''  # 当前题目类型
    
    for para in doc.paragraphs:
        text = para.text.strip()  # 去除首尾空白
        if not text:
            continue
            
        # 处理题型标记（如 ##单选题##）
        if text.startswith('##') and text.endswith('##'):
            question_type = text.strip('#').strip()
            continue
            
        # 处理题目（以数字开头的段落）- 修改正则表达式以适应更多格式
        if re.match(r'^[#]?\d+\.?\s', text):  # 添加可选的点号
            if current_question:  # 如果存在上一题，保存它
                questions.append(current_question)
            # 初始化新题目
            current_question = {
                'type': question_type,
                'question': '',
                'options': [],
                'answer': ''
            }
            
            # 提取题目内容和答案
            question_text = re.sub(r'^[#]?\d+\.?\s', '', text).strip()  # 移除题号和可能的点号
            # 提取答案（在{}中的内容）
            answer_match = re.search(r'\{(.+?)\}', question_text)
            if answer_match:
                current_question['answer'] = answer_match.group(1)
                question_text = re.sub(r'\{.+?\}', '', question_text)  # 移除答案标记
            current_question['question'] = question_text.strip()
            
        # 处理选项（A-D开头的行）- 修改以适应带点号的格式
        elif re.match(r'^[A-D]\.?\s', text):  # 添加可选的点号
            option_text = re.sub(r'^[A-D]\.?\s', '', text).strip()  # 移除选项标记和可能的点号
            current_question['options'].append(option_text)
    
    # 保存最后一题
    if current_question:
        questions.append(current_question)
    
    return questions

def determine_question_type(question):
    """根据答案和选项特征判断题型"""
    answer = question['answer'].strip().upper()
    options = question['options']
    
    # 判断题特征：答案为"正确"/"错误"或"对"/"错"
    if answer in ['正确', '错误', '对', '错', 'T', 'F']:
        return '判断题'
    
    # 多选题特征：答案包含多个字母
    if len(answer) > 1 and all(c in 'ABCDEFGHIJK' for c in answer):
        return '多选题'
    
    # 单选题特征：答案为单个字母A-K
    if len(answer) == 1 and answer in 'ABCDEFGHIJK':
        return '单选题'
    
    # 其他情况默认为单选题
    return '单选题'

def write_to_excel(questions, output):
    """
    将题目写入Excel文件
    参数:
        questions: 题目列表
        output: 可以是文件路径字符串或BytesIO对象
    """
    wb = Workbook()
    ws = wb.active
    current_row = 1
    
    # 获取所有选项字母
    def get_option_letters(questions):
        letters = set()
        for q in questions:
            for i, _ in enumerate(q['options']):
                letters.add(chr(65 + i))  # 将数字转换为对应的字母(A=65, B=66, ...)
        return sorted(list(letters))
    
    # 动态生成表头
    def get_headers(question_type, option_letters):
        common_headers = ['题型', '题目']
        if question_type in ['单选题', '多选题']:
            return common_headers + option_letters + ['答案']
        elif question_type == '判断题':
            return common_headers + ['正确', '错误', '答案']
        else:
            return common_headers + ['答案']
    
    # 确定所有可能的选项字母
    option_letters = get_option_letters(questions)
    
    # 处理每个问题，确定题型
    processed_questions = []
    for q in questions:
        if not q.get('type'):
            q['type'] = determine_question_type(q)
        processed_questions.append(q)
    
    # 写入表头
    headers = get_headers(processed_questions[0]['type'], option_letters)
    for col, header in enumerate(headers, 1):
        ws.cell(row=current_row, column=col, value=header)
    current_row += 1
    
    # 写入题目内容
    for q in processed_questions:
        question_type = q['type']
        ws.cell(row=current_row, column=1, value=question_type)
        ws.cell(row=current_row, column=2, value=q['question'])
        
        if question_type in ['单选题', '多选题']:
            # 动态处理选择题选项
            for i, option in enumerate(q['options']):
                col_index = headers.index(chr(65 + i)) + 1
                ws.cell(row=current_row, column=col_index, value=option)
            ws.cell(row=current_row, column=len(headers), value=q['answer'])
            
        elif question_type == '判断题':
            # 处理判断题选项
            ws.cell(row=current_row, column=3, value='正确')
            ws.cell(row=current_row, column=4, value='错误')
            ws.cell(row=current_row, column=5, value=q['answer'])
            
        else:
            # 其他题型只记录答案
            ws.cell(row=current_row, column=3, value=q['answer'])
        
        current_row += 1
    
    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width
    
    # 根据output类型选择保存方式
    if isinstance(output, str):
        wb.save(output)
    else:
        wb.save(output)  # BytesIO对象

def main():
    """主函数：处理用户界面和程序流程"""
    # 设置页面标题
    st.title("Word试题转Excel工具")
    st.write("支持的题型：单选题、多选题、判断题")
    
    # 显示使用说明（可折叠）
    with st.expander("查看使用说明"):
        st.markdown("""
        ### 使用说明：
        1. 上传Word文档（.docx格式）
        2. Word文档格式要求：
           - 题型标记必须使用 ##单选题##、##多选题##、##判断题## 格式（包含前后的##号）
           - 每道题必须以数字或者#开头（如：#1. 或 1、）
           - 答案必须放在题目末尾的大括号中，如：这是一道题{A}
           - 选项必须以A、B、C、D开头（可以使用A. 或 A、格式）
           - 判断题答案格式：{正确}、{错误}、{对}、{错}、{T}、{F}
           - 多选题答案格式：{ABC}（多个选项直接相连）
           - 单选题答案格式：{A}（单个选项）
        
        3. 保存文件：
           - 输入完整的保存路径（如：D:/Documents 或 C:/Users/Username/Desktop）
           - 输入文件名（无需添加.xlsx后缀）
           - 点击"保存Excel文件"按钮完成保存
        
        ### 示例题目格式：
        ```
        ##单选题##
        1. 这是一道单选题{A}
        A. 选项一
        B. 选项二
        C. 选项三
        D. 选项四
        
        ##多选题##
        2. 这是一道多选题{ABC}
        A. 选项一
        B. 选项二
        C. 选项三
        D. 选项四
        
        ##判断题##
        3. 这是一道判断题{正确}
        ```
        """)
    
    # 文件上传组件
    uploaded_file = st.file_uploader("选择Word文档", type=['docx'])
    
    # 处理上传的文件
    if uploaded_file:
        try:
            # 读取和处理Word文档
            doc = Document(uploaded_file)
            questions = extract_questions_from_docx(doc)
            
            if questions:
                # 添加保存方式选择
                save_method = st.radio(
                    "选择保存方式：",
                    ["直接下载", "指定保存位置"]
                )
                
                save_filename = st.text_input("文件名（不需要.xlsx后缀）：", value="转换结果")
                
                if save_method == "直接下载":
                    # 创建Excel文件的字节流
                    excel_buffer = io.BytesIO()
                    wb = Workbook()
                    write_to_excel(questions, excel_buffer)
                    excel_buffer.seek(0)
                    
                    # 添加下载按钮
                    st.download_button(
                        label="下载Excel文件",
                        data=excel_buffer,
                        file_name=f"{save_filename}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                else:
                    # 手动指定保存位置
                    save_path = st.text_input("保存路径（如：D:/Documents）：", value="")
                    if save_path:
                        full_path = f"{save_path.rstrip('/')}/{save_filename}.xlsx"
                        
                        if st.button("保存Excel文件"):
                            try:
                                os.makedirs(os.path.dirname(full_path), exist_ok=True)
                                write_to_excel(questions, full_path)
                                st.success(f"文件已成功保存至：{full_path}")
                            except Exception as e:
                                st.error(f"保存文件时出错：{str(e)}")
                
                # 显示转换结果预览
                st.subheader("转换结果预览")
                for i, q in enumerate(questions, 1):
                    st.write(f"**{i}. {q['type']}**")
                    st.write(f"题目：{q['question']}")
                    if q['options']:
                        st.write("选项：")
                        for j, opt in enumerate(q['options']):
                            st.write(f"{chr(65+j)}. {opt}")
                    st.write(f"答案：{q['answer']}")
                    st.write("---")
                
        except Exception as e:
            st.error(f"处理文件时出错：{str(e)}")
                
# 程序入口点
if __name__ == '__main__':
    main()